import json
import logging
import os
import pickle
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from config import APP_NAMES, DIMESHIFT_NAME
from global_log import GlobalLog
from utils.file_utils import get_coverage_targets_file
from utils.stats_utils import (
    vargha_delaney,
    wilcoxon_unpaired_pvalue,
)

import argparse

args = argparse.ArgumentParser()
args.add_argument(
    "--app-name",
    help="Name of the app under test",
    type=str,
    choices=APP_NAMES,
    default=DIMESHIFT_NAME,
)
args.add_argument(
    "--results-folder-name",
    help="Name of the results folder",
    type=str,
    default="results",
)
args.add_argument(
    "--raw-coverage-targets",
    help="Consider raw coverage targets instead of coverage percentage",
    action="store_true",
    default=False,
)
args.add_argument(
    "--percentage-iterations",
    help="Percentage of iterations to consider to compute efficiency (AUC)",
    type=int,
    default=50,
)

args, _ = args.parse_known_args()

if __name__ == "__main__":

    app_name = args.app_name
    raw_coverage_targets = args.raw_coverage_targets
    results_folder_name = args.results_folder_name
    percentage_iterations = args.percentage_iterations
    font_size = 30
    line_width = 4
    fig_size = (16, 9)
    plot_legend = False
    plot_coverage_zoom = True
    coverage_zoom_percentage_max_length = 50
    BASELINE = "random"

    assert percentage_iterations in np.arange(
        0, 100 + 1
    ), "Number of iterations be between 0 and 100"

    results_dir = os.path.join(results_folder_name, app_name)

    assert os.path.exists(results_dir), f"Results for {app_name} do not exist"

    alpha_significance_level = 0.05
    analysis_dir = os.path.join(results_dir, "analysis")
    os.makedirs(analysis_dir, exist_ok=True)

    logger = GlobalLog("analyze")

    # redirect stdout to a file
    logging.basicConfig(
        filename=os.path.join(analysis_dir, "analyze.log"),
        level=logging.INFO,
        filemode="w",
    )

    techniques = dict()

    coverage_targets_file = get_coverage_targets_file(app_name=app_name)
    with open(coverage_targets_file, "rb") as f:
        coverage_targets = set(pickle.load(f))

    for technique in os.listdir(results_dir):
        if technique == "analysis":
            continue

        if technique not in techniques:
            techniques[technique] = dict()

        coverages_over_time = []
        global_coverages = []
        global_raw_coverages = []
        lengths_over_time = []
        num_tests = []
        avg_execution_times = []
        avg_generation_times = []
        for json_file in filter(
            lambda x: x.endswith(".json"),
            os.listdir(os.path.join(results_dir, technique)),
        ):
            with open(os.path.join(results_dir, technique, json_file), "r") as f:
                data = json.load(f)
                if raw_coverage_targets:
                    coverage_over_time = list(
                        map(
                            lambda x: x * len(coverage_targets) / 100,
                            data["coverage_over_time"],
                        )
                    )
                else:
                    coverage_over_time = data["coverage_over_time"]

                coverage_target_file = f"{json_file.split('.')[0]}_covered_targets.pkl"
                with open(
                    os.path.join(results_dir, technique, coverage_target_file), "rb"
                ) as f2:
                    covered_targets = set(pickle.load(f2))
                    global_raw_coverages.append(len(set(covered_targets)))

                coverages_over_time.append(coverage_over_time)
                global_coverages.append(data["global_coverage"])
                lengths_over_time.append(data["individual_lengths"])
                num_tests.append(len(coverage_over_time))
                avg_execution_times.append(data["avg_execution_time"])
                avg_generation_times.append(data["avg_generation_time"])

        techniques[technique]["coverages_over_time"] = coverages_over_time
        techniques[technique]["global_coverages"] = global_coverages
        techniques[technique]["global_raw_coverages"] = global_raw_coverages
        techniques[technique]["lengths_over_time"] = lengths_over_time
        techniques[technique]["num_tests"] = num_tests
        techniques[technique]["avg_execution_times"] = avg_execution_times
        techniques[technique]["avg_generation_times"] = avg_generation_times

    max_lengths = [
        max([len(c) for c in techniques[technique]["coverages_over_time"]])
        for technique in techniques.keys()
    ]
    max_length = max(max_lengths)

    min_lengths = [
        min([len(c) for c in techniques[technique]["coverages_over_time"]])
        for technique in techniques.keys()
    ]
    min_length = min(min_lengths)

    efficiencies = dict()
    efficiences_at = dict()
    coverages = dict()
    num_tests = dict()
    lengths = dict()
    avg_execution_times = dict()
    avg_generation_times = dict()

    # area of a rectangle = base * height
    max_area = (
        max_length * len(coverage_targets) if raw_coverage_targets else max_length * 100
    )

    plt.figure(figsize=fig_size)
    plt.rcParams["font.size"] = font_size
    plt.rcParams["font.weight"] = "bold"
    for technique in list(techniques.keys()):
        coverages_over_time = techniques[technique]["coverages_over_time"]
        for i in range(len(coverages_over_time)):
            coverages_over_time[i] = np.pad(
                coverages_over_time[i],
                (0, max_length - len(coverages_over_time[i])),
                mode="edge",
            )

        coverages_at = []
        for i, c in enumerate(coverages_over_time):
            num_iterations = percentage_iterations * len(c) // 100
            assert (
                len(c) >= num_iterations
            ), f"There are less than {num_iterations} iterations in technique {technique} at run {i} (corresponding to {percentage_iterations}%)"
            coverages_at.append(c[:num_iterations])

        # area under the curve
        areas = np.trapz(coverages_over_time, axis=1) / max_area
        efficiencies[technique] = areas.tolist()

        # area under the curve at N iterations
        areas_at = np.trapz(coverages_at, axis=1) / max_area
        efficiences_at[technique] = areas_at.tolist()

        coverages_over_time = np.array(coverages_over_time)
        mean_coverage = np.mean(coverages_over_time, axis=0)
        std_mean_coverage = np.std(coverages_over_time, axis=0) / np.sqrt(
            len(coverages_over_time)
        )

        if raw_coverage_targets:
            coverages[technique] = techniques[technique]["global_raw_coverages"]
        else:
            coverages[technique] = techniques[technique]["global_coverages"]

        num_tests[technique] = techniques[technique]["num_tests"]
        lengths[technique] = techniques[technique]["lengths_over_time"]
        avg_execution_times[technique] = techniques[technique]["avg_execution_times"]
        avg_generation_times[technique] = techniques[technique]["avg_generation_times"]

        plt.plot(mean_coverage, label=technique, linewidth=line_width)
        plt.fill_between(
            range(len(mean_coverage)),
            mean_coverage - std_mean_coverage,
            mean_coverage + std_mean_coverage,
            alpha=0.3,
        )

    vs = "_".join(list(techniques.keys()))

    plt.xlabel("# Executed Tests", fontsize=font_size, fontweight="bold")
    if raw_coverage_targets:
        plt.ylabel("Coverage (# targets)", fontsize=font_size, fontweight="bold")
    else:
        plt.ylabel("Coverage (%)", fontsize=font_size, fontweight="bold")

    if plot_legend:
        plt.legend()
    plt.title(f"{app_name}", fontsize=font_size, fontweight="bold")
    plt.savefig(
        os.path.join(analysis_dir, f"coverage_over_time_{vs}.svg"), bbox_inches="tight"
    )

    if plot_coverage_zoom:
        plt.figure(figsize=fig_size)
        plt.rcParams["font.size"] = font_size
        plt.rcParams["font.weight"] = "bold"
        min_iterations = np.inf
        for technique in list(techniques.keys()):
            coverages_over_time = techniques[technique]["coverages_over_time"]
            for i in range(len(coverages_over_time)):
                coverages_over_time[i] = np.pad(
                    coverages_over_time[i],
                    (0, max_length - len(coverages_over_time[i])),
                    mode="edge",
                )

            coverages_zoom = []
            for i, c in enumerate(coverages_over_time):
                num_iterations = coverage_zoom_percentage_max_length * len(c) // 100
                min_iterations = min(min_iterations, num_iterations)
                assert (
                    len(c) >= num_iterations
                ), f"There are less than {num_iterations} iterations in technique {technique} at run {i} (corresponding to {percentage_iterations}%)"
                coverages_zoom.append(c[num_iterations:])

            coverages_over_time = np.array(coverages_zoom)
            mean_coverage = np.mean(coverages_over_time, axis=0)
            std_mean_coverage = np.std(coverages_over_time, axis=0) / np.sqrt(
                len(coverages_over_time)
            )

            plt.plot(
                np.arange(num_iterations, max_length),
                mean_coverage,
                label=technique,
                linewidth=line_width + 2,
            )
            plt.fill_between(
                np.arange(num_iterations, max_length),
                mean_coverage - std_mean_coverage,
                mean_coverage + std_mean_coverage,
                alpha=0.2,
            )

        vs = "_".join(list(techniques.keys()))

        plt.xlabel("# Executed Tests", fontsize=font_size, fontweight="bold")
        plt.xlim(min_iterations, max_length)
        if raw_coverage_targets:
            plt.ylabel("Coverage (# targets)", fontsize=font_size, fontweight="bold")
        else:
            plt.ylabel("Coverage (%)", fontsize=font_size, fontweight="bold")

        if plot_legend:
            plt.legend()
        plt.title(f"{app_name}", fontsize=font_size, fontweight="bold")
        plt.savefig(
            os.path.join(analysis_dir, f"coverage_over_time_zoom_{vs}.svg"),
            bbox_inches="tight",
        )

    plt.figure(figsize=fig_size)
    plt.rcParams["font.size"] = font_size
    plt.rcParams["font.weight"] = "bold"
    window_size = 100
    for technique in list(techniques.keys()):
        lengths_over_time = techniques[technique]["lengths_over_time"]
        for i in range(len(lengths_over_time)):

            lengths_over_time[i] = lengths_over_time[i][:min_length]

            # lengths_over_time[i] = np.pad(
            #     lengths_over_time[i],
            #     (0, max_length - len(lengths_over_time[i])),
            #     mode="symmetric",
            # )

            kernel = np.ones(window_size) / window_size

            # # smoothing over a window of window_size
            lengths_over_time[i] = np.convolve(
                lengths_over_time[i],
                kernel,
                mode="valid",
            )

        mean_length = np.mean(lengths_over_time, axis=0)
        std_error_mean_length = np.std(lengths_over_time, axis=0) / np.sqrt(
            len(lengths_over_time)
        )
        plt.plot(mean_length, label=technique, linewidth=line_width)
        plt.fill_between(
            range(len(mean_length)),
            mean_length - std_error_mean_length,
            mean_length + std_error_mean_length,
            alpha=0.3,
        )

    vs = "_".join(list(techniques.keys()))

    plt.xlabel("# Executed Tests", fontsize=font_size, fontweight="bold")
    plt.ylabel("Lenght (# Statements)", fontsize=font_size, fontweight="bold")
    if plot_legend:
        plt.legend()
    plt.title(f"{app_name}", fontsize=font_size, fontweight="bold")
    plt.rcParams["font.size"] = 14
    plt.rcParams["font.weight"] = "bold"
    plt.savefig(
        os.path.join(analysis_dir, f"length_over_time_{vs}.svg"),
        bbox_inches="tight",
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    rqs_sheet = workbook.create_sheet("RQs")
    other_metrics_sheet = workbook.create_sheet("Others")

    letters = "BCDEFGHIJKLMNOPQRSTUVWXYZ"
    assert len(techniques.keys()) <= len(letters), "Too many techniques to display"

    rqs_sheet["A1"] = app_name
    other_metrics_sheet["A1"] = app_name

    list_keys_techniques = list(techniques.keys())
    assert (
        BASELINE in list_keys_techniques
    ), f"{BASELINE} should be in the list of techniques"
    list_keys_techniques.remove(BASELINE)
    list_keys_techniques.sort()
    list_keys_techniques.insert(0, BASELINE)

    j = 0
    k = 0

    for i, technique in enumerate(list_keys_techniques):

        mean_coverage = np.mean(coverages[technique])
        std_mean_coverage = np.std(coverages[technique])
        logger.info(
            f"{technique} mean coverage: {mean_coverage:.2f} ± {std_mean_coverage:.2f}"
        )

        mean_efficiency = np.mean(efficiencies[technique])
        std_efficiency = np.std(efficiencies[technique])
        logger.info(
            f"{technique} mean efficiency: {mean_efficiency:.2f} ± {std_efficiency:.2f}"
        )

        mean_efficiency_at = np.mean(efficiences_at[technique])
        std_efficiency_at = np.std(efficiences_at[technique])
        logger.info(
            f"{technique} mean efficiency @ {percentage_iterations}% iterations: {mean_efficiency_at:.2f} ± {std_efficiency_at:.2f}"
        )

        assert len(letters) > j + 2, "Too many techniques to display"
        rqs_sheet.merge_cells(f"{letters[j]}1:{letters[j + 2]}1")
        rqs_sheet[f"{letters[j]}1"] = technique
        rqs_sheet[f"{letters[j]}1"].alignment = Alignment(horizontal="center")

        rqs_sheet[f"{letters[j]}2"] = (
            "Raw Coverage" if raw_coverage_targets else "Coverage (%)"
        )
        rqs_sheet[f"{letters[j]}2"].alignment = Alignment(horizontal="center")
        rqs_sheet[f"{letters[j + 1]}2"] = "AUC (%)"
        rqs_sheet[f"{letters[j + 1]}2"].alignment = Alignment(horizontal="center")
        rqs_sheet[f"{letters[j + 2]}2"] = f"AUC@{percentage_iterations}% (%)"
        rqs_sheet[f"{letters[j + 2]}2"].alignment = Alignment(horizontal="center")

        rqs_sheet[f"{letters[j]}3"] = mean_coverage
        rqs_sheet[f"{letters[j]}3"].alignment = Alignment(horizontal="center")
        rqs_sheet[f"{letters[j + 1]}3"] = mean_efficiency
        rqs_sheet[f"{letters[j + 1]}3"].alignment = Alignment(horizontal="center")
        rqs_sheet[f"{letters[j + 2]}3"] = mean_efficiency_at
        rqs_sheet[f"{letters[j + 2]}3"].alignment = Alignment(horizontal="center")

        assert len(letters) > k + 3, "Too many techniques to display"

        mean_num_tests = np.mean(num_tests[technique])
        std_num_tests = np.std(num_tests[technique])
        logger.info(
            f"{technique} num tests: {mean_num_tests:.2f} ± {std_num_tests:.2f}"
        )

        mean_length = np.mean(lengths[technique])
        std_length = np.std(lengths[technique])
        logger.info(f"{technique} mean length: {mean_length:.2f} ± {std_length:.2f}")

        mean_avg_execution_time = np.mean(avg_execution_times[technique])
        std_avg_execution_time = np.std(avg_execution_times[technique])
        logger.info(
            f"{technique} avg execution time: {mean_avg_execution_time:.2f} ± {std_avg_execution_time:.2f}"
        )

        mean_avg_generation_time = np.mean(avg_generation_times[technique])
        std_avg_generation_time = np.std(avg_generation_times[technique])
        logger.info(
            f"{technique} avg generation time: {mean_avg_generation_time:.2f} ± {std_avg_generation_time:.2f}"
        )

        other_metrics_sheet.merge_cells(f"{letters[k]}1:{letters[k + 3]}1")
        other_metrics_sheet[f"{letters[k]}1"] = technique
        other_metrics_sheet[f"{letters[k]}1"].alignment = Alignment(horizontal="center")

        other_metrics_sheet[f"{letters[k]}2"] = "# Executed Tests"
        other_metrics_sheet[f"{letters[k]}2"].alignment = Alignment(horizontal="center")
        other_metrics_sheet[f"{letters[k + 1]}2"] = "Length"
        other_metrics_sheet[f"{letters[k + 1]}2"].alignment = Alignment(
            horizontal="center"
        )
        other_metrics_sheet[f"{letters[k + 2]}2"] = "Avg Execution Time (s)"
        other_metrics_sheet[f"{letters[k + 2]}2"].alignment = Alignment(
            horizontal="center"
        )
        other_metrics_sheet[f"{letters[k + 3]}2"] = "Avg Generation Time (s)"
        other_metrics_sheet[f"{letters[k + 3]}2"].alignment = Alignment(
            horizontal="center"
        )

        other_metrics_sheet[f"{letters[k]}3"] = mean_num_tests
        other_metrics_sheet[f"{letters[k]}3"].alignment = Alignment(horizontal="center")
        other_metrics_sheet[f"{letters[k + 1]}3"] = mean_length
        other_metrics_sheet[f"{letters[k + 1]}3"].alignment = Alignment(
            horizontal="center"
        )
        other_metrics_sheet[f"{letters[k + 2]}3"] = mean_avg_execution_time
        other_metrics_sheet[f"{letters[k + 2]}3"].alignment = Alignment(
            horizontal="center"
        )
        other_metrics_sheet[f"{letters[k + 3]}3"] = mean_avg_generation_time
        other_metrics_sheet[f"{letters[k + 3]}3"].alignment = Alignment(
            horizontal="center"
        )

        j += 3
        k += 4

        logger.info("")

    workbook.save(filename=os.path.join(analysis_dir, "table_metrics.xlsx"))

    workbook = Workbook()
    workbook.remove(workbook.active)
    coverage_sheet = workbook.create_sheet("Coverage")
    efficiency_sheet = workbook.create_sheet("Efficiency")
    efficiency_at_sheet = workbook.create_sheet(
        f"Efficiency @ {percentage_iterations}%"
    )
    coverage_sheet["A1"] = app_name
    efficiency_sheet["A1"] = app_name
    efficiency_at_sheet["A1"] = app_name

    logger.info("=" * 5 + " Statistical Analysis " + "=" * 5)

    for i in range(len(list_keys_techniques)):
        technique1 = list_keys_techniques[i]
        if i != len(list_keys_techniques) - 1:
            coverage_sheet[f"A{i + 2}"] = technique1
            efficiency_sheet[f"A{i + 2}"] = technique1
            efficiency_at_sheet[f"A{i + 2}"] = technique1
        for j in range(i + 1, len(list_keys_techniques)):

            technique2 = list_keys_techniques[j]
            if i == 0:
                coverage_sheet[f"{letters[j - 1]}{i + 1}"] = technique2
                efficiency_sheet[f"{letters[j - 1]}{i + 1}"] = technique2
                efficiency_at_sheet[f"{letters[j - 1]}{i + 1}"] = technique2

            logger.info(f"* {technique1} vs {technique2} *")

            p_value = wilcoxon_unpaired_pvalue(
                a=coverages[technique1], b=coverages[technique2]
            )
            logger.info(f"Coverage p-value {technique1} vs {technique2}: {p_value:.2f}")
            p_value_effect_size_coverage = f"{p_value:.2f}"
            if p_value < alpha_significance_level:
                estimate, magnitude = vargha_delaney(
                    a=coverages[technique1], b=coverages[technique2]
                )
                logger.info(
                    f"Coverage effect size {technique1} vs {technique2}: {estimate:.2f} ({magnitude})"
                )
                p_value_effect_size_coverage += f"_({estimate:.2f} {magnitude})"

            p_value = wilcoxon_unpaired_pvalue(
                a=efficiencies[technique1], b=efficiencies[technique2]
            )
            logger.info(
                f"Efficiency p-value {technique1} vs {technique2}: {p_value:.2f}"
            )
            p_value_effect_size_efficiency = f"{p_value:.2f}"
            if p_value < alpha_significance_level:
                estimate, magnitude = vargha_delaney(
                    a=efficiencies[technique1], b=efficiencies[technique2]
                )
                logger.info(
                    f"Efficiency effect size {technique1} vs {technique2}: {estimate:.2f} ({magnitude})"
                )
                p_value_effect_size_efficiency += f"_({estimate:.2f} {magnitude})"

            p_value = wilcoxon_unpaired_pvalue(
                a=efficiences_at[technique1], b=efficiences_at[technique2]
            )
            logger.info(
                f"Efficiency @ {percentage_iterations}% p-value {technique1} vs {technique2}: {p_value:.2f}"
            )
            p_value_effect_size_efficiency_at = f"{p_value:.2f}"
            if p_value < alpha_significance_level:
                estimate, magnitude = vargha_delaney(
                    a=efficiences_at[technique1], b=efficiences_at[technique2]
                )
                logger.info(
                    f"Efficiency @ {percentage_iterations}% effect size {technique1} vs {technique2}: {estimate:.2f} ({magnitude})"
                )
                p_value_effect_size_efficiency_at += f"_({estimate:.2f} {magnitude})"

            coverage_sheet[f"{letters[j - 1]}{i + 2}"] = (
                f"{p_value_effect_size_coverage}"
            )
            efficiency_sheet[f"{letters[j - 1]}{i + 2}"] = (
                f"{p_value_effect_size_efficiency}"
            )
            efficiency_at_sheet[f"{letters[j - 1]}{i + 2}"] = (
                f"{p_value_effect_size_efficiency_at}"
            )

            if i < len(list(techniques.keys())) - 1:
                logger.info("-" * 10)

    workbook.save(filename=os.path.join(analysis_dir, "table_metrics_statistics.xlsx"))
