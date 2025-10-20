'''
帮我写个本程序功能说明：
1. 解析命令行参数，获取应用名称和结果文件夹名称。
2. 创建分析目录，用于存储分析结果。
3. 遍历结果目录中的测试方法，加载覆盖目标数据。
4. 计算困难目标统计，包括困难目标数量、困难目标覆盖率和困难目标平均覆盖率。
5. 分析独特目标，统计每个测试方法覆盖的独特目标数量。
6. 将分析结果写入Excel文件，包括困难目标统计和独特目标分析。
7. 执行统计显著性分析，比较不同测试方法之间的困难目标覆盖率。
8. 将统计显著性分析结果写入Excel文件。
'''

from collections import defaultdict
import json
import logging
import os
import pickle

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment

from config import APP_NAMES, DIMESHIFT_NAME
from global_log import GlobalLog
from utils.file_utils import get_coverage_targets_file

import argparse

from utils.stats_utils import vargha_delaney, wilcoxon_unpaired_pvalue

# 命令行参数解析
args = argparse.ArgumentParser()
args.add_argument(
    "--app-name",
    help="Name of the app under test",
    type=str,
    choices=APP_NAMES,
    default=DIMESHIFT_NAME,
)
args.add_argument(
    "--results-folder-name",
    help="Name of the results folder",
    type=str,
    default="results",
)

args, _ = args.parse_known_args()


if __name__ == "__main__":

    # 获取命令行参数
    app_name = args.app_name
    results_folder_name = args.results_folder_name
    BASELINE = "random"  # 基线方法

    alpha_significance_level = 0.05  # 显著性水平

    techniques = dict()  # 存储不同测试方法的数据
    num_runs = []  # 运行次数列表
    results_dir = os.path.join(results_folder_name, app_name)

    # 创建分析目录
    analysis_dir = os.path.join(results_dir, "analysis")
    os.makedirs(analysis_dir, exist_ok=True)

    logger = GlobalLog("analyze_coverage")

    # 重定向标准输出到文件
    logging.basicConfig(
        filename=os.path.join(analysis_dir, "analyze_coverage.log"),
        level=logging.INFO,
        filemode="w",
    )

    # 遍历结果目录中的测试方法
    for technique in os.listdir(results_dir):
        if technique == "analysis":
            continue

        if technique not in techniques:
            techniques[technique] = dict()

        # 加载覆盖目标数据
        covered_targets_technique = []
        for pkl_file in filter(
            lambda x: x.endswith(".pkl") and x.endswith("covered_targets.pkl"),
            os.listdir(os.path.join(results_dir, technique)),
        ):
            with open(os.path.join(results_dir, technique, pkl_file), "rb") as f:
                covered_targets = pickle.load(f)
                covered_targets_technique.append(covered_targets)

        techniques[technique]["covered_targets"] = covered_targets_technique

        # 加载测试数量数据
        num_tests = []
        for json_file in filter(
            lambda x: x.endswith(".json"),
            os.listdir(os.path.join(results_dir, technique)),
        ):
            with open(os.path.join(results_dir, technique, json_file), "r") as f:
                data = json.load(f)
                num_tests.append(len(data["coverage_over_time"]))

        techniques[technique]["num_tests"] = num_tests
        num_runs.append(len(techniques[technique]["covered_targets"]))

    # 检查所有方法的运行次数是否相同
    assert (
        len(set(num_runs)) == 1
    ), "Number of runs should be the same for all techniques"

    # 加载覆盖目标文件
    coverage_targets_file = get_coverage_targets_file(app_name=app_name)
    with open(coverage_targets_file, "rb") as f:
        coverage_targets = set(pickle.load(f))

    uncovered_targets = set([ct.clone() for ct in coverage_targets])
    all_targets = set([ct.clone() for ct in coverage_targets])

    logger.info(f"All targets: {len(all_targets)}")

    # 创建Excel工作簿
    workbook = Workbook()
    workbook.remove(workbook.active)
    complexity_sheet = workbook.create_sheet("Complexity")

    letters = "BCDEFGHIJKLMNOPQRSTUVWXYZ"
    assert len(techniques.keys()) <= len(letters), "Too many techniques to display"

    complexity_sheet["A1"] = app_name
    unique_targets = defaultdict(list)

    # 基线方法的覆盖目标分析
    covered_targets_baseline = {ct: [0] * num_runs[0] for ct in all_targets}
    covered_targets_baseline_cp_estimation = {
        ct: [0] * num_runs[0] for ct in all_targets
    }
    for i in range(num_runs[0]):
        assert (
            BASELINE in techniques.keys()
        ), f"{BASELINE} should be in the list of techniques"
        covered_targets_baseline_i = techniques[BASELINE]["covered_targets"][i]
        for ct in covered_targets_baseline_i:
            covered_targets_baseline[ct][i] = 1
            covered_targets_baseline_cp_estimation[ct][i] = techniques[BASELINE][
                "num_tests"
            ][i] / (ct.iteration + 1)

    # 计算困难目标的统计
    probability_threshold = 0.0
    coverage_probability_threshold = 10**-3
    num_difficult_targets = 0
    num_difficult_targets_cp_estimation = 0
    for ct in covered_targets_baseline.keys():
        coverage_probability = np.mean(covered_targets_baseline[ct])
        coverage_probability_estimation = (
            np.mean(covered_targets_baseline_cp_estimation[ct])
            if sum(covered_targets_baseline_cp_estimation[ct]) > 0
            else 0.0
        )
        logger.info(
            f"Target {ct.target_type}-{ct.line_number}: {covered_targets_baseline[ct]}, coverage probability: {coverage_probability}"
        )
        logger.info(
            f"Target {ct.target_type}-{ct.line_number}, coverage probability estimation: {coverage_probability_estimation}"
        )
        if coverage_probability <= probability_threshold:
            num_difficult_targets += 1
        if coverage_probability_estimation <= coverage_probability_threshold:
            num_difficult_targets_cp_estimation += 1

    logger.info(
        f"Num targets with coverage probability by random < {probability_threshold}: {num_difficult_targets} ({num_difficult_targets/len(all_targets)*100:.2f} %)"
    )
    logger.info(
        f"Num targets with coverage probability estimation < {coverage_probability_threshold}: {num_difficult_targets_cp_estimation} ({num_difficult_targets_cp_estimation/len(all_targets)*100:.2f} %)"
    )

    # 分析每个运行的独特目标
    for i in range(num_runs[0]):
        logger.info(f"========== Run {i} ==========")
        for technique_1 in techniques.keys():
            covered_targets_technique_1 = techniques[technique_1]["covered_targets"][i]
            uncovered_targets = uncovered_targets - set(covered_targets_technique_1)
            uncovered_targets_technique_1 = all_targets - set(
                covered_targets_technique_1
            )
            common_targets = set()
            for technique_2 in techniques.keys():
                if technique_1 == technique_2:
                    continue
                covered_targets_technique_2 = techniques[technique_2][
                    "covered_targets"
                ][i]
                common_targets_tech_1_tech_2 = list(
                    filter(
                        lambda x: x in covered_targets_technique_1,
                        covered_targets_technique_2,
                    )
                )
                common_targets.update(common_targets_tech_1_tech_2)
            unique_targets_technique_1 = (
                set(covered_targets_technique_1) - common_targets
            )
            unique_targets[technique_1].append(len(unique_targets_technique_1))
            logger.info(
                f"Unique targets {technique_1}: {len(unique_targets_technique_1)}"
            )
            logger.info(
                f"Unique targets: {[ct.__str__() for ct in unique_targets_technique_1]}"
            )
            logger.info(
                f"Uncovered targets {technique_1}: {[ct.__str__() for ct in uncovered_targets_technique_1]}"
            )

    # 排序技术列表，基线方法排在最前面
    list_keys_techniques = list(techniques.keys())
    assert (
        BASELINE in list_keys_techniques
    ), f"{BASELINE} should be in the list of techniques"
    list_keys_techniques.remove(BASELINE)
    list_keys_techniques.sort()
    list_keys_techniques.insert(0, BASELINE)

    # 将独特目标数据写入Excel
    j = 0
    for i, technique in enumerate(list_keys_techniques):
        assert len(letters) > j + 1, "Too many techniques to display"
        complexity_sheet[f"{letters[j]}1"] = technique
        complexity_sheet[f"{letters[j]}1"].alignment = Alignment(horizontal="center")

        complexity_sheet[f"{letters[j]}2"] = "Unique Targets"
        complexity_sheet[f"{letters[j]}2"].alignment = Alignment(horizontal="center")

        complexity_sheet[f"{letters[j]}3"] = (
            np.mean(unique_targets[technique])
            if sum(unique_targets[technique]) > 0
            else 0.0
        )
        complexity_sheet[f"{letters[j]}3"].alignment = Alignment(horizontal="center")

        j += 1

    workbook.save(filename=os.path.join(analysis_dir, "table_complexity.xlsx"))

    logger.info(
        f"Targets not covered by any technique: {len(uncovered_targets)}, {[ct.__str__() for ct in uncovered_targets]}"
    )

    # 创建统计分析的Excel文件
    workbook = Workbook()
    workbook.remove(workbook.active)
    unique_targets_sheet = workbook.create_sheet("Complexity")
    unique_targets_sheet["A1"] = app_name

    logger.info("=" * 5 + " Statistical Analysis " + "=" * 5)

    # 执行统计显著性分析
    for i in range(len(list_keys_techniques)):
        technique1 = list_keys_techniques[i]
        if i != len(list_keys_techniques) - 1:
            unique_targets_sheet[f"A{i + 2}"] = technique1
        for j in range(i + 1, len(list_keys_techniques)):

            technique2 = list_keys_techniques[j]
            if i == 0:
                unique_targets_sheet[f"{letters[j - 1]}{i + 1}"] = technique2

            logger.info(f"* {technique1} vs {technique2} *")

            # 计算p值和效应大小
            p_value = wilcoxon_unpaired_pvalue(
                a=unique_targets[technique1], b=unique_targets[technique2]
            )
            logger.info(
                f"Unique targets p-value {technique1} ({np.mean(unique_targets[technique1])}) vs {technique2} ({np.mean(unique_targets[technique2])}): {p_value:.2f}"
            )
            p_value_effect_size_unique_targets = f"{p_value:.2f}"
            if p_value < alpha_significance_level:
                estimate, magnitude = vargha_delaney(
                    a=unique_targets[technique1], b=unique_targets[technique2]
                )
                logger.info(
                    f"Unique targets effect size {technique1} ({np.mean(unique_targets[technique1])}) vs {technique2} ({np.mean(unique_targets[technique2])}): {estimate:.2f} ({magnitude})"
                )
                p_value_effect_size_unique_targets += f"_({estimate:.2f} {magnitude})"

            unique_targets_sheet[f"{letters[j - 1]}{i + 2}"] = (
                f"{p_value_effect_size_unique_targets}"
            )

            if i < len(list(techniques.keys())) - 1:
                logger.info("-" * 10)

    workbook.save(
        filename=os.path.join(analysis_dir, "table_complexity_statistics.xlsx")
    )

    # 分析每个方法的最困难目标
    logger.info("========== Most difficult target for each technique ==========")
    for technique in techniques.keys():
        logger.info(f"Technique {technique}:")
        difficult_targets = []
        for i in range(len(techniques[technique])):
            covered_targets_technique = techniques[technique]["covered_targets"][i]
            covered_targets_technique = sorted(
                covered_targets_technique, key=lambda x: x.iteration, reverse=True
            )
            for j in range(3):
                ct = covered_targets_technique[j].clone()
                difficult_targets.append(ct)

        difficult_targets = sorted(
            difficult_targets, key=lambda x: x.iteration, reverse=True
        )
        logger.info(
            f"Targets: {[difficult_target.__str__() for difficult_target in difficult_targets]}"
        )
        with open(
            os.path.join(analysis_dir, f"{technique}_difficult_targets.txt"), "w"
        ) as f:
            for difficult_target in difficult_targets:
                f.write(f"{difficult_target.__str__()}\n")
